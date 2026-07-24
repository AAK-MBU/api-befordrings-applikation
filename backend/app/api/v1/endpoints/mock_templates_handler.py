"""
╔══════════════════════════════════════════════════════════════════════╗
║ 🔥 TEMPORARY MOCK - REMOVE WHEN api-skabelonmotor IS LIVE 🔥          ║
╠══════════════════════════════════════════════════════════════════════╣
║ Self-contained, in-process copy of the                                ║
║ `/templates_handler/update_template_data/{process}` endpoint from the ║
║ api-skabelonmotor service (app/api/templates_handler.py +             ║
║ app/utils/helper_functions.parse_workbook_afgoerelsesbrev).           ║
║                                                                        ║
║ It exists ONLY so the "Opdater skabelondata" button can work before   ║
║ api-skabelonmotor is dockerised and online. It fetches the source     ║
║ template files from SharePoint, parses them, and upserts the result   ║
║ into the rpa.Templates table that the afgoerelsesbreve RPA reads.      ║
║                                                                        ║
║ When the API is live: delete this file, remove its include_router in  ║
║ app/api/v1/api.py, drop the temporary mbu_msoffice_integration dep in ║
║ pyproject.toml, and point the endpoint at the real API instead (a     ║
║ thin requests.get to skabelonmotor/templates_handler/...).            ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import json
import os
import re
import urllib.parse

from io import BytesIO

from fastapi import APIRouter, HTTPException

from mbu_msoffice_integration.sharepoint_class import Sharepoint

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from sqlalchemy import create_engine, text


router = APIRouter(prefix="/templates_handler", tags=["Templates handler (TEMP mock)"])

# Regex used to detect block headers such as "Blok 1", "Blok 3.1", "Blok 7.2a"
BLOCK_HEADER_PATTERN = re.compile(r"^Blok\s+([0-9]+(?:\.\s*[0-9]+)?[a-zA-Z]?)")

# SharePoint credentials - same env vars used by api-skabelonmotor
SHAREPOINT_KWARGS = {
    "tenant": os.getenv("TENANT"),
    "client_id": os.getenv("CLIENT_ID"),
    "thumbprint": os.getenv("APPREG_THUMBPRINT"),
    "cert_path": os.getenv("GRAPH_CERT_PEM"),
}


@router.get("/update_template_data/{process}")
def update_template_data(process: str):
    """
    Refresh the DB copy of a process' template data from the SharePoint source
    files. Fetches the Word template + Excel workbook, parses the workbook into
    JSON, and upserts both into rpa.Templates.
    """

    if process != "afgoerelsesbreve":
        raise HTTPException(
            status_code=400,
            detail=f"Ukendt proces: {process} - skabelondata blev ikke opdateret!",
        )

    sharepoint = Sharepoint(
        site_url="https://aarhuskommune.sharepoint.com/",
        site_name="MBURPA",
        document_library="Delte dokumenter",
        **SHAREPOINT_KWARGS,
    )

    # Sharepoint._auth() swallows auth errors and leaves ctx as None. Surface it
    # instead of failing later with a confusing download/parse error.
    if sharepoint.ctx is None:
        raise HTTPException(
            status_code=502,
            detail="SharePoint-godkendelse fejlede (tjek TENANT/CLIENT_ID/"
                   "APPREG_THUMBPRINT/GRAPH_CERT_PEM og TLS-opsætning i containeren).",
        )

    folder_name = "Egenbefordring/Afgørelsesbreve"

    template_binary_docx = sharepoint.fetch_file_using_open_binary(
        file_name="skabelon.docx",
        folder_name=folder_name,
    )
    _require_ooxml(template_binary_docx, folder_name, "skabelon.docx")

    binary_excel = sharepoint.fetch_file_using_open_binary(
        file_name="Afgørelsesbreve.xlsm",
        folder_name=folder_name,
    )
    _require_ooxml(binary_excel, folder_name, "Afgørelsesbreve.xlsm")

    json_data = parse_workbook_afgoerelsesbrev(binary_excel=binary_excel)

    query = """
        BEGIN TRANSACTION;

        UPDATE rpa.Templates
        SET
            word_template = :word_template,
            workbook_json = :workbook_json,
            last_updated = SYSDATETIME()
        WHERE process_name = :process_name;

        IF @@ROWCOUNT = 0
        BEGIN
            INSERT INTO rpa.Templates (
                process_name,
                word_template,
                workbook_json
            )
            VALUES (
                :process_name,
                :word_template,
                :workbook_json
            );
        END

        COMMIT;
    """

    params = {
        "process_name": process,
        "word_template": template_binary_docx,
        "workbook_json": json.dumps(json_data),
    }

    execute_sql(
        query=query,
        params=params,
        conn_string=os.getenv("DBCONNECTIONSTRINGPROD"),
    )

    return {"message": "Skabelondata blev succesfuldt opdateret."}


def _require_ooxml(binary: bytes | None, folder_name: str, file_name: str) -> None:
    """
    Validate that a SharePoint download is a real Office (OOXML) file.

    .docx / .xlsm are ZIP containers, so their bytes start with the ZIP magic
    number "PK\\x03\\x04". fetch_file_using_open_binary does NOT raise on a 404 -
    it returns the error response body - so without this check a missing/mis-
    named file would flow into load_workbook and fail with a confusing
    "File is not a zip file". Here we surface what actually came back instead.
    """

    path = f"{folder_name}/{file_name}"

    if not binary:
        raise HTTPException(
            status_code=502,
            detail=f"Kunne ikke hente '{path}' fra SharePoint (tomt svar - "
                   f"tjek at filen findes og at sti/navn er korrekt).",
        )

    if not binary.startswith(b"PK\x03\x04"):
        # Not a zip -> almost certainly a SharePoint error body. Show a snippet.
        snippet = binary[:300].decode("utf-8", errors="replace").strip()
        raise HTTPException(
            status_code=502,
            detail=f"'{path}' var ikke en gyldig Office-fil. SharePoint svarede: {snippet}",
        )


def execute_sql(query: str, params: dict, conn_string: str) -> int:
    """
    Run an INSERT/UPDATE/DELETE sql statement against the rpa.Templates database.
    """

    encoded_conn_str = urllib.parse.quote_plus(conn_string)

    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={encoded_conn_str}")

    with engine.begin() as conn:
        result = conn.execute(text(query), params)

    return result.rowcount


def extract_cell_formatting(cell):
    """
    Convert Excel rich text content into HTML-like formatted text.
    """

    if cell is None or cell.value is None:
        return ""

    value = cell.value

    # ----------------------------------------
    # Rich formatted text (Excel rich text)
    # ----------------------------------------
    if isinstance(value, CellRichText):

        parts = []

        for block in value:

            if isinstance(block, str):
                text_part = block
                font = None
            else:
                text_part = block.text or ""
                font = block.font

            if not text_part:
                continue

            # Remove zero-width characters sometimes inserted by Excel
            text_part = text_part.replace("​", "")

            # Replace Excel tab indentation
            text_part = text_part.replace("\t", " ")

            prefix = ""
            suffix = ""

            if font:

                # Bold
                if font.b:
                    prefix += "<strong>"
                    suffix = "</strong>" + suffix

                # Italic
                if font.i:
                    prefix += "<em>"
                    suffix = "</em>" + suffix

                # Underline
                if font.u in ["single", "double", "singleAccounting", "doubleAccounting", True]:
                    prefix += "<u>"
                    suffix = "</u>" + suffix

                # Strikethrough
                if font.strike:
                    prefix += "<strike>"
                    suffix = "</strike>" + suffix

                # Text color
                if font.color and font.color.rgb:
                    rgb = font.color.rgb[-6:]

                    # Skip default black text
                    if rgb != "000000":
                        prefix += f'<span style="color:#{rgb}">'
                        suffix = "</span>" + suffix

            parts.append(f"{prefix}{text_part}{suffix}")

        return "".join(parts)

    # ----------------------------------------
    # Plain text cell (no formatting)
    # ----------------------------------------
    return str(value)


def parse_workbook_afgoerelsesbrev(binary_excel: bytes) -> list[dict]:
    """
    Pure Excel parser.

    Extracts blocks and their entries from the workbook without applying
    any business logic, metadata, or custom functions.
    """

    LINK_MAPPING = {
        "Folkeskoleloven (retsinformation.dk)": "https://www.retsinformation.dk/eli/lta/2025/1100#P26",

        "Bekendtgørelse om befordring af elever i folkeskolen (retsinformation.dk)": "https://www.retsinformation.dk/eli/lta/2014/688",

        "Ungdomsskoleloven (retsinformation.dk)": "https://www.retsinformation.dk/eli/lta/2019/608",

        "Bekendtgørelse af lov om befordringsrabat til uddannelsessøgende i ungdomsuddannelser m.v": "https://www.retsinformation.dk/eli/lta/2026/379#P10",

        "Behandling af personoplysninger i Børn og Unge (aarhus.dk)": "https://aarhus.dk/om-kommunen/databeskyttelse/behandling-af-personoplysninger-i-boern-og-unge"
    }

    def inject_links(entry_text: str) -> str:
        for link_text, url in LINK_MAPPING.items():
            if link_text in entry_text:
                entry_text = entry_text.replace(
                    link_text,
                    f'<a href="{url}">{link_text}</a>'
                )

        return entry_text

    wb = load_workbook(BytesIO(binary_excel), rich_text=True)

    parsed_blocks = []
    current_block = None

    # ----------------------------------------
    # Parse workbook sheets
    # ----------------------------------------
    for sheet_name in wb.sheetnames:

        if not sheet_name.startswith("Blok"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows())

        for i, row in enumerate(rows):
            col_a_cell = row[0] if len(row) > 0 else None
            col_b_cell = row[1] if len(row) > 1 else None

            col_a = col_a_cell.value if col_a_cell else None
            col_b = extract_cell_formatting(col_b_cell) if col_b_cell else None

            # ----------------------------------------
            # Detect block header
            # ----------------------------------------
            if isinstance(col_a, str):

                match = BLOCK_HEADER_PATTERN.match(col_a)

                if match:

                    block_id = match.group(1).replace(" ", "").strip()

                    # Mapping key from column C in next row
                    next_row = rows[i + 1] if i + 1 < len(rows) else None
                    next_col_c = None

                    if next_row and len(next_row) > 2:
                        next_col_c = next_row[2].value

                    current_block = {
                        "block_id": block_id,
                        "title": col_a,
                        "mapping": str(next_col_c).strip() if next_col_c else None,
                        "entries": {}
                    }

                    parsed_blocks.append(current_block)

                    continue

            if not current_block:
                continue

            # ----------------------------------------
            # Parse entries
            # ----------------------------------------
            if col_a and col_b:

                entry_text = col_b.strip()

                # Skip "Ingen tekst"
                if normalize_key(entry_text) == "ingentekst":
                    continue

                key = str(col_a)

                entry_text = inject_links(entry_text)

                current_block["entries"][key] = entry_text

    return parsed_blocks


def normalize_key(value: str) -> str:
    """
    Normalize strings for reliable key comparisons.
    """

    return (
        value.strip()
        .lower()
        .replace(" ", "")
        .replace(".", "")
        .replace("ø", "oe")
        .replace("å", "aa")
        .replace("æ", "ae")
        .replace("?", "")
        .replace("-", "")
        .replace("_", "")
    )
